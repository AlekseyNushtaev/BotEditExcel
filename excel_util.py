import openpyxl
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

def open_exel_1(filepath):
    wb_obj = openpyxl.load_workbook(filepath)
    sheet_obj = wb_obj['Sheet1']

    dct = {}
    length = sheet_obj.max_row

    for i in range(1, length):
        dct[i] = sheet_obj.cell(row=i + 1, column=3).value
    return dct

def open_exel_2(filepath, dct):
    wb_obj = openpyxl.load_workbook(filepath, data_only=True)
    sheet_obj = wb_obj['ИТОГО']

    res = {}
    length = sheet_obj.max_row
    for i in range(8, len(dct.keys()) + 8):
        lst = [dct[i - 7]]
        for y in range(4, length):
            quantity = sheet_obj.cell(row=y, column=i).value
            if quantity not in [0, '0', None]:
                lst.append([str(sheet_obj.cell(row=y, column=6).value), quantity])
        res[i-7] = lst
    return res

def create_excel(res, filepath):
    wb_obj = openpyxl.Workbook()
    ss_sheet = wb_obj['Sheet']
    ss_sheet.title = 'Sheet1'
    sheet_obj = wb_obj.active
    cell = sheet_obj.cell(row=1, column=1)
    cell.value = "баркод товара"
    cell = sheet_obj.cell(row=1, column=2)
    cell.value = "кол-во товаров"
    cell = sheet_obj.cell(row=1, column=3)
    cell.value = "шк короба"
    cell = sheet_obj.cell(row=1, column=4)
    cell.value = "срок годности"
    count = 2
    for i in range(1, len(res.keys()) + 1):
        for y in range(len(res[i]) - 1):
            cell = sheet_obj.cell(row=count, column=1)
            cell.value = res[i][y+1][0]
            cell = sheet_obj.cell(row=count, column=2)
            cell.value = res[i][y + 1][1]
            cell = sheet_obj.cell(row=count, column=3)
            cell.value = res[i][0]
            count += 1
    wb_obj.save(filepath)

